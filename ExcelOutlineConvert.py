import os

import keyboard
from openpyxl import load_workbook
from openpyxl import Workbook


class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


def ParsingXLSX(in_file):
    print(f"{bcolors.OKBLUE}=======================================")
    print(f"{bcolors.OKBLUE}DEF ParsingXLSX - {in_file}")
    workbook = load_workbook(filename=in_file)
    sheet = workbook.active
    print(f"{bcolors.OKGREEN}File Count: row = {sheet.max_row}, column = {sheet.max_column}")

    # Read from space Row
    firstin = True
    u_maxrow = sheet.max_row
    for i in range(1, int(sheet.max_row)):
        if firstin and not sheet.cell(row=i, column=1).value:
            firstin = False
            u_maxrow = i

    # Read from space Column
    firstin = True
    u_maxcol = sheet.max_column
    for i in range(1, int(sheet.max_column)):
        if firstin and not sheet.cell(row=1, column=i).value:
            firstin = False
            u_maxcol = i

    print(f"{bcolors.OKGREEN}User Count: row = {u_maxrow}, column = {u_maxcol}")

    # Read XLSX to Array
    for i in range(1, u_maxrow):
        cur_livel = sheet.row_dimensions[i].outlineLevel
        row = {}
        row["Livel"] = cur_livel
        row["Row"] = i
        for j in range(1, u_maxcol+1):
            cells = sheet.cell(row=i, column=j).value if sheet.cell(row=i, column=j).value else ''
            row[f"Col{j}"] = str(cells).strip()
        print(f"{bcolors.FAIL}    {i}: {cur_livel} = {cells}; {row}")
        DataSetXLSX.append(row)


def ProcessGroupToLine():
    print(f"{bcolors.OKBLUE}=======================================")
    print(f"{bcolors.OKBLUE}DEF ProcessGroupToLine")

    # Maximum livel
    max_livel = 0
    for row in DataSetXLSX:
        cur_livel = row.get('Livel')
        max_livel = cur_livel if cur_livel > max_livel else max_livel  # max livel
    print(f"{bcolors.OKGREEN}Max Livel: {max_livel}")

    # Main
    livel_keys = {}
    for i in range(0, max_livel + 1):
        livel_keys[i] = ""

    first_step = True
    pre_livel = 0

    for row in DataSetXLSX:
        cur_livel = int(row.get('Livel'))
        cur_row = int(row.get('Row'))
        cur_key = row.get('Col1')
        cur_name = row.get('Col3')

        if first_step:
            first_step = False
            pre_livel = cur_livel
            print(f"{bcolors.OKGREEN}First: {cur_livel}")
            livel_keys[cur_livel] = cur_key
        else:
            if cur_livel > pre_livel:
                livel_keys[cur_livel] = row.get('Col1')
                print(f"{bcolors.FAIL}+Row {cur_row}: LVL={cur_livel}; KEY={cur_key}/{cur_name}; TREE[{livel_keys}]")
                pre_livel = cur_livel
            else:

                if cur_livel == pre_livel:
                    print(
                        f"{bcolors.FAIL}=Row {cur_row}: LVL={cur_livel}; KEY={cur_key}/{cur_name}; TREE[{livel_keys}]")
                    livel_keys[cur_livel] = cur_key
                else:
                    if cur_livel < pre_livel:
                        livel_keys[pre_livel] = ""
                        livel_keys[cur_livel] = cur_key
                        print(
                            f"{bcolors.FAIL}-Row {cur_row}: LVL={cur_livel}; KEY={cur_key}/{cur_name}; TREE[{livel_keys}]")
                        pre_livel = cur_livel

        insrow = {}
        insrow['Row'] = str(cur_row) if cur_row != 1 else 'Row'
        insrow['Livel'] = str(cur_livel) if cur_row != 1 else 'Livel'

        UpGroup = ''
        LastGr = True
        for i in range(0, max_livel + 1):
            insrow[f"GR_{i}"] = livel_keys.get(i) if cur_row != 1 else f"GR_{i}"
            if livel_keys.get(i):
                UpGroup = livel_keys.get((i - 1)) if i>0 else ''
            # Prev group
            # if LastGr and livel_keys.get(i) == '':
            #     UpGroup = livel_keys.get((i - 1))
            #     LastGr = False

        insrow[f"UpGroup"] = UpGroup if cur_row != 1 else 'UpGroup'
        for el in row:
            if el != 'Row' and el != 'Livel':
                insrow[el] = row.get(el)

        DataSetProcess.append(insrow)


def CreateOutFile(in_file):
    print(f"{bcolors.OKBLUE}=======================================")
    print(f"{bcolors.OKBLUE}DEF CreateOutFile")
    filename, file_extension = os.path.splitext(in_file)
    out_file = f"{filename}_new{file_extension}"
    print(f"{bcolors.OKGREEN}Output: {out_file}")

    # Delete out file if exist
    if os.path.isfile(out_file):
        os.remove(out_file)
        print(f"{bcolors.OKGREEN}Delete exist Out file")

    # Create book
    print(f"{bcolors.OKGREEN}Create book")
    book = Workbook()
    sheet = book.active
    exRow = 0
    for row in DataSetProcess:
        exRow += 1
        exCol = 0
        sRow = ""
        for el in row:
            exCol += 1
            sheet.cell(row=exRow, column=exCol).value = row.get(el) if row.get(el) else ""
            sRow += f"{el} = {row.get(el)}; "
        print(f"{bcolors.FAIL}{sRow}")
    print(f"{bcolors.OKGREEN}Save book")
    book.save(out_file)


def main():
    for in_file in os.listdir(os.getcwd()):
        if os.path.isfile(in_file) and in_file.endswith(".xlsx") and in_file.find("_new") == -1:
            # Input file
            print(f"{bcolors.OKGREEN}Input: {in_file}")

            # Parse file
            DataSetXLSX.clear()  # clear dataset
            ParsingXLSX(f"{os.getcwd()}\\{in_file}")
            print(f"{bcolors.OKGREEN}{DataSetXLSX}")

            # Group to Line
            DataSetProcess.clear()  # clear dataset
            ProcessGroupToLine()
            print(f"{bcolors.OKGREEN}{DataSetProcess}")

            # Out file
            CreateOutFile(f"{os.getcwd()}\\{in_file}")


if __name__ == '__main__':
    print(f"{bcolors.HEADER}Create: Cherepanov Maxim masygreen@gmail.com (c), 11.2020")
    DataSetXLSX = []
    DataSetProcess = []
    main()
    print(f'{bcolors.HEADER}\n\n*All Process done. \n*Press Space to Exit ... It the longest shortcut \_(o0)_\...')
    keyboard.wait("space")
